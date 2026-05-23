# src/cee_comps/output_excel.py
"""Excel comp-set deliverable (live-formula model).

Reads `enriched_comps`, writes three sheets:
  - Cover:    title, author, date, methodology, sources, FX stamp
  - Comps:    input block (hardcoded EUR values, blue) + multiples
              and per-bucket mean/median as live Excel formulas
  - Profiles: one-line descriptions + full LLM classifications

Design: raw EUR inputs (EV, revenue, EBITDA) are hardcoded; every
multiple and every summary stat is an Excel formula referencing them,
so a reviewer can change an input and the sheet recomputes. Formulas
are IFERROR-wrapped so missing-data rows (e.g. Comarch) render "n/m"
rather than #DIV/0!.

openpyxl writes formulas as strings with no cached values, but
`wb.calculation.fullCalcOnLoad = True` (set in `build()`) flags the
workbook so Excel and LibreOffice recompute every formula on open —
no manual Ctrl+F9 / save-as needed.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rich.console import Console

from . import db
from .config import EXCEL_OUT, FX_DATE
from .schemas import REVENUE_MODEL_LABELS, REVENUE_MODEL_ORDER

console = Console()

FONT_NAME = "Arial"
BLUE = "0000FF"      # hardcoded inputs (xlsx skill convention)
BLACK = "000000"     # formulas
HEADER_FILL = "1F3864"
HEADER_FONT = "FFFFFF"
GROUP_FILL = "D6E4F0"

MULT_FMT = "0.0x"
PCT_FMT = "0.0%"
EUR_FMT = "#,##0;(#,##0);-"
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(cell):
    cell.font = Font(name=FONT_NAME, bold=True, color=HEADER_FONT)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX


def _build_cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "CEE Software & IT Services — Comparable Companies"
    ws["B2"].font = Font(name=FONT_NAME, size=18, bold=True)
    ws["B4"] = "Tomasz Szymula"
    ws["B5"] = f"FX as of {FX_DATE} · all monetary values in EUR millions"
    ws["B5"].font = Font(name=FONT_NAME, italic=True, color="595959")

    ws["B7"] = "Methodology"
    ws["B7"].font = Font(name=FONT_NAME, size=12, bold=True)
    method = (
        "Market data sourced from Yahoo Finance (yfinance); Polish names "
        "with incomplete coverage backfilled from most recent annual "
        "reports. Business-model classifications extracted from annual "
        "report excerpts via Claude with Pydantic-validated structured "
        "outputs. Multiples computed on an EUR-normalised basis at a "
        f"single spot rate dated {FX_DATE}. EV/EBITDA shown as 'n/m' "
        "where LTM EBITDA is non-positive and excluded from peer "
        "statistics. Multiples and summary statistics are live Excel "
        "formulas over the hardcoded input block on the Comps sheet."
    )
    ws["B8"] = method
    ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B8:H14")

    ws["B16"] = "Data sources"
    ws["B16"].font = Font(name=FONT_NAME, size=12, bold=True)
    ws["B17"] = "Market data: Yahoo Finance · Classifications: Claude (Anthropic)"
    ws["B18"] = "Fallback financials: company annual reports (see README)"

    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16


def _build_comps(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Comps")
    ws.sheet_view.showGridLines = False

    cols = [
        ("company_name", "Company", None),
        ("ticker", "Ticker", None),
        ("currency", "Ccy", None),
        ("enterprise_value_eur", "EV (\u20acmm)", EUR_FMT),
        ("ltm_revenue_eur", "LTM Rev (\u20acmm)", EUR_FMT),
        ("ltm_ebitda_eur", "LTM EBITDA (\u20acmm)", EUR_FMT),
        ("ev_revenue", "EV/Rev", MULT_FMT),
        ("ev_ebitda", "EV/EBITDA", MULT_FMT),
        ("revenue_growth_yoy", "Rev Growth", PCT_FMT),
        ("ebitda_margin", "EBITDA Mgn", PCT_FMT),
        ("rule_of_40", "Rule of 40", "0.0"),
    ]
    # Column letters (for formulas) and 1-based indices (for ws.cell) keyed
    # by field name. Built once so the per-row writes stay readable.
    C = {key: get_column_letter(i + 1) for i, (key, _, _) in enumerate(cols)}
    COL_IDX = {key: i + 1 for i, (key, _, _) in enumerate(cols)}

    # Header
    for i, (_, label, _) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        _hdr(c)
    ws.freeze_panes = "A2"

    r = 2
    group_stat_rows: dict[str, list[int]] = {}
    for model in REVENUE_MODEL_ORDER:
        block = df[df["revenue_model"] == model]
        if block.empty:
            continue

        # Group banner
        ws.cell(row=r, column=1, value=REVENUE_MODEL_LABELS.get(model, model))
        for i in range(1, len(cols) + 1):
            cc = ws.cell(row=r, column=i)
            cc.fill = PatternFill("solid", fgColor=GROUP_FILL)
            cc.font = Font(name=FONT_NAME, bold=True)
        r += 1

        first_data = r
        for _, row in block.iterrows():
            ev, rev, eb = C["enterprise_value_eur"], C["ltm_revenue_eur"], C["ltm_ebitda_eur"]
            ws.cell(row=r, column=1, value=row["company_name"]).font = Font(name=FONT_NAME)
            ws.cell(row=r, column=2, value=row["ticker"]).font = Font(name=FONT_NAME)
            ws.cell(row=r, column=3, value=row["currency"]).font = Font(name=FONT_NAME)

            # Hardcoded EUR inputs — blue, per xlsx skill convention
            for key in ("enterprise_value_eur", "ltm_revenue_eur", "ltm_ebitda_eur"):
                val = row[key]
                cell = ws.cell(
                    row=r, column=COL_IDX[key],
                    value=None if pd.isna(val) else float(val),
                )
                cell.font = Font(name=FONT_NAME, color=BLUE)
                cell.number_format = EUR_FMT

            # Live formulas — black
            ws.cell(row=r, column=7, value=f"=IFERROR({ev}{r}/{rev}{r},\"n/m\")")
            ws.cell(row=r, column=8,
                    value=f"=IF({eb}{r}<=0,\"n/m\",IFERROR({ev}{r}/{eb}{r},\"n/m\"))")
            gv = row["revenue_growth_yoy"]
            # ISBLANK guard: a blank EBITDA cell would otherwise evaluate as
            # 0 in Excel arithmetic and silently render the margin as 0.0%.
            mg = f'=IF(ISBLANK({eb}{r}),"n/m",IFERROR({eb}{r}/{rev}{r},"n/m"))'
            # Revenue growth is a hardcoded input (from yfinance), so it gets
            # the blue "input" treatment alongside EV / Revenue / EBITDA.
            growth_cell = ws.cell(row=r, column=9,
                                  value=None if pd.isna(gv) else float(gv))
            growth_cell.font = Font(name=FONT_NAME, color=BLUE)
            growth_cell.number_format = PCT_FMT
            ws.cell(row=r, column=10, value=mg).number_format = PCT_FMT
            ws.cell(row=r, column=11,
                    value=(f"=IFERROR(({C['revenue_growth_yoy']}{r}"
                           f"+{C['ebitda_margin']}{r})*100,\"n/m\")"))
            for i in (7, 8, 11):
                ws.cell(row=r, column=i).font = Font(name=FONT_NAME, color=BLACK)
            ws.cell(row=r, column=7).number_format = MULT_FMT
            ws.cell(row=r, column=8).number_format = MULT_FMT
            ws.cell(row=r, column=11).number_format = "0.0"
            r += 1
        last_data = r - 1

        # Mean / median rows — formulas over the group's data range only.
        # Text "n/m" cells are ignored by AVERAGE/MEDIAN automatically.
        for stat in ("AVERAGE", "MEDIAN"):
            ws.cell(row=r, column=1,
                    value=f"{REVENUE_MODEL_LABELS.get(model, model)} — {stat.title()}")
            ws.cell(row=r, column=1).font = Font(name=FONT_NAME, italic=True, bold=True)
            for i in (7, 8, 9, 10, 11):
                col = get_column_letter(i)
                cell = ws.cell(
                    row=r, column=i,
                    value=f"=IFERROR({stat}({col}{first_data}:{col}{last_data}),\"\")",
                )
                cell.font = Font(name=FONT_NAME, italic=True, bold=True, color=BLACK)
                cell.number_format = (
                    PCT_FMT if i in (9, 10) else "0.0" if i == 11 else MULT_FMT
                )
            group_stat_rows.setdefault(model, []).append(r)
            r += 1
        r += 1  # spacer row between groups

    # Conditional 3-colour scale on EV/Rev and EV/EBITDA data cells
    last_row = r
    for col_letter in (C["ev_revenue"], C["ev_ebitda"]):
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    ws.column_dimensions["A"].width = 26
    for i in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


def _build_profiles(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Profiles")
    ws.sheet_view.showGridLines = False
    fields = [
        ("company_name", "Company"),
        ("ticker", "Ticker"),
        ("one_line_description", "Description"),
        ("revenue_model", "Revenue Model"),
        ("end_market", "End Market"),
        ("customer_profile", "Customer"),
        ("primary_geography", "Geography"),
        ("rd_intensity_signal", "R&D Intensity"),
    ]
    for i, (_, label) in enumerate(fields, start=1):
        _hdr(ws.cell(row=1, column=i, value=label))
    ws.freeze_panes = "A2"

    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for i, (key, _) in enumerate(fields, start=1):
            v = row.get(key)
            cell = ws.cell(row=r, column=i, value="" if pd.isna(v) else str(v))
            cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(
                wrap_text=(key == "one_line_description"), vertical="top"
            )

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["C"].width = 60
    for col in "BDEFGH":
        ws.column_dimensions[col].width = 16


def build() -> None:
    """Read enriched_comps and write the Excel deliverable."""
    df = db.read_table("enriched_comps")

    # Sort once here so Comps and Profiles share the same row order — a
    # reviewer cross-referencing between sheets shouldn't have to hunt.
    # Revenue-model bucket first (per REVENUE_MODEL_ORDER), then ascending
    # EV/Revenue within each bucket.
    df = df.copy()
    df["_order"] = df["revenue_model"].map(
        {m: i for i, m in enumerate(REVENUE_MODEL_ORDER)}
    ).fillna(99)
    df = df.sort_values(["_order", "ev_revenue"], na_position="last")

    wb = Workbook()
    # Force a full recalc on open. openpyxl writes formulas without
    # cached values; without this flag, Excel/LibreOffice would render
    # the formula cells as blanks until the user manually saved the file.
    wb.calculation.fullCalcOnLoad = True

    _build_cover(wb)
    _build_comps(wb, df)
    _build_profiles(wb, df)
    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_OUT)
    console.print(
        f"[green]Wrote {EXCEL_OUT.name} ({len(df)} companies, 3 sheets).[/green]"
    )